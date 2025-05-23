import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import uuid

# File path
EXCEL_FILE = "Accounts 2025 V1.0 - Sample.xlsm"
TX_SHEET = "Tx"
BALANCE_SHEET = "Balance"
SALARY_SHEET = "SLRY(US)"

# Load data from Excel
@st.cache_data(show_spinner=False)
def load_data():
    tx = pd.read_excel(EXCEL_FILE, sheet_name=TX_SHEET, engine="openpyxl")
    balance = pd.read_excel(EXCEL_FILE, sheet_name=BALANCE_SHEET, engine="openpyxl")
    resources = pd.read_excel(EXCEL_FILE, sheet_name="Resources", engine="openpyxl")
    # Fetch account names from Balance sheet rows A6 to A30
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[BALANCE_SHEET]
    account_names = [ws[f"A{row}"].value for row in range(6, 31) if ws[f"A{row}"].value]
    return tx, balance, resources, account_names

def generate_unique_id(account, amount):
    today = datetime.today().strftime("%Y%m%d")
    return f"{today}.{account.lower().replace(' ', '')}.{amount:.2f}"[:30]

# Append new transaction to Excel
def add_transaction_to_excel(entry):
    wb = load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb[TX_SHEET]

    # Identify the last non-empty row based on column A (unique_id)
    next_row = ws.max_row + 1
    for row_num in range(ws.max_row, 0, -1):
        if ws.cell(row=row_num, column=1).value not in [None, ""]:
            next_row = row_num + 1
            break

    for row in entry:
        for col, value in enumerate(row, start=1):
            ws.cell(row=next_row, column=col, value=value)
        next_row += 1

    # Sort rows by date (column B) in descending order
    max_row = ws.max_row
    ws.auto_filter.ref = None  # Clear existing filters
    data_range = f"A2:M{max_row}"
    ws.auto_filter.ref = data_range
    ws.sort_range(
        sort_by_columns=[(2, False)],  # Column B, descending
        sort_orientation="rows",
        header=False
    )

    wb.save(EXCEL_FILE)

# UI Start
st.set_page_config(page_title="Jaunty Minty", layout="wide")
st.title("🌿 Jaunty Minty: Your Personal Finance Companion")

# Tabs for navigation
tab1, tab2, tab3, tab4 = st.tabs(["➕ Add 🍃 Transaction", "📄 💸 Transactions", "📊 💼 Balances", "💼 🧾 Salary"])

with tab1:
    st.subheader("Record a Fresh Transaction ✨")

    _, _, resources_df, account_names = load_data()
    category_options = resources_df['Category'].dropna().unique().tolist() if 'Category' in resources_df.columns else []
    billed_where_options = resources_df['Billed Where'].dropna().unique().tolist() if 'Billed Where' in resources_df.columns else []

    col1, col2 = st.columns(2)
    with col1:
        tx_date = st.date_input("Date", value=datetime.today())
        tx_type = st.radio("Type", ["Expense", "Income", "Transfer"])
        category_options_clean = sorted(set(category_options))
        category = st.selectbox(
            "Category",
            options=["", *category_options_clean, "Add new..."],
            index=0,
            key="category_select"
        )
    if category == "Add new...":
        new_category = st.text_input("Enter new category", key="new_category")
        if new_category:
            category = new_category
            if st.button("Add New Category", key="add_category_btn"):
                wb = load_workbook(EXCEL_FILE, keep_vba=True)
                ws = wb["Resources"]
                ws.append([new_category])
                wb.save(EXCEL_FILE)
                st.success(f"Added new category: {new_category}")
                st.experimental_rerun()
    if category and category not in category_options:
        if st.button("Add New Category"):
            wb = load_workbook(EXCEL_FILE, keep_vba=True)
            ws = wb["Resources"]
            ws.append([category])
            wb.save(EXCEL_FILE)
            st.success(f"Added new category: {category}")
        account_options = ["", *account_names]
        account = st.selectbox("Account", options=account_options, index=0)
        amount = st.number_input("Amount", min_value=0.0, step=0.01)
        comments = st.text_input("Comments")

    with col2:
        expand_bill = st.checkbox("Expanded Bill?")
        billed_where_options_clean = sorted(set(billed_where_options))
        billed_where = st.selectbox(
            "Billed Where",
            options=["", *billed_where_options_clean, "Add new..."],
            index=0,
            key="billed_where_select"
        )
    if billed_where == "Add new...":
        new_billed_where = st.text_input("Enter new Billed Where", key="new_billed_where")
        if new_billed_where:
            billed_where = new_billed_where
            if st.button("Add New Billed Where", key="add_billed_where_btn"):
                wb = load_workbook(EXCEL_FILE, keep_vba=True)
                ws = wb["Resources"]
                ws.append([None, new_billed_where])
                wb.save(EXCEL_FILE)
                st.success(f"Added new Billed Where: {new_billed_where}")
                st.experimental_rerun()
    if billed_where and billed_where not in billed_where_options:
        if st.button("Add New Billed Where"):
            wb = load_workbook(EXCEL_FILE, keep_vba=True)
            ws = wb["Resources"]
            row = [None, billed_where]
            ws.append(row)
            wb.save(EXCEL_FILE)
            st.success(f"Added new Billed Where: {billed_where}")
        currency = st.selectbox("Currency", ["USD", "INR"])

    bill_items = []
    if expand_bill:
        st.markdown("---")
        st.markdown("### Bill Breakdown")
        bill_count = st.number_input("Number of Items", min_value=1, step=1, value=1)

        for i in range(bill_count):
            st.markdown(f"**Item {i+1}**")
            col_a, col_b, col_c = st.columns([1, 1, 2])
            with col_a:
                item_amount = st.number_input(f"Amount {i+1}", min_value=0.0, step=0.01, key=f"amt_{i}")
            with col_b:
                item_category_options = sorted(set(category_options))
                item_category = st.selectbox(
                    f"Category {i+1}",
                    options=["", *item_category_options, "Add new..."],
                    key=f"cat_{i}_select"
                )
            if item_category == "Add new...":
                item_category = st.text_input(f"New Category {i+1}", key=f"cat_{i}_new")
            with col_c:
                item_comment = st.text_input(f"Comment {i+1}", key=f"cmt_{i}")
            bill_items.append({"amount": item_amount, "category": item_category, "comment": item_comment})

    if st.button("Submit Transaction"):
        if tx_type != "Transfer" and expand_bill and billed_where.strip() == "":
            st.error("Expand Bill is checked, but Billed Where is empty.")
        elif category.strip() == "" and not expand_bill:
            st.error("Category field is required.")
        elif expand_bill and round(sum(i["amount"] for i in bill_items), 2) != round(amount, 2):
            st.error("Bill amount does not tally with breakdown.")
        else:
            entries = []
            unique_id = generate_unique_id(account, amount)

            # Main entry (non-expanded)
            main_entry = [
                unique_id, tx_date, tx_type, tx_type, billed_where, False,
                category, account, amount, amount, currency, comments, None
            ]
            entries.append(main_entry)

            # Expanded items
            if expand_bill:
                for item in bill_items:
                    sub_entry = [
                        unique_id, tx_date, tx_type, tx_type, billed_where, True,
                        item["category"], account, None, item["amount"], currency,
                        item["comment"], None
                    ]
                    entries.append(sub_entry)

            add_transaction_to_excel(entries)
            st.success("Transaction saved successfully!")

with tab2:
    st.subheader("View All Transactions 🧾")
    tx_df, _, _, _ = load_data()
    st.dataframe(tx_df.sort_values(by="date", ascending=False), use_container_width=True)

with tab3:
    st.subheader("Your Minty Balance Sheet 💹")
    _, balance_df, _, _ = load_data()
    st.markdown("### 🧾 Overall Balance")
    st.dataframe(balance_df.iloc[0:3], use_container_width=True)
    st.markdown("### 🏦 Account-wise Balances")
    st.dataframe(balance_df.iloc[5:30], use_container_width=True)

with tab4:
    st.subheader("Log Your Salary 🌱")
    salary_date = st.date_input("Salary Date", key="slry")
    col1, col2 = st.columns(2)
    with col1:
        acct_options = ["", *account_names]
        acct1 = st.selectbox("Account 1", options=acct_options, index=0)
        amt1 = st.number_input("Amount 1", key="amt1")
    with col2:
        acct2 = st.selectbox("Account 2", options=acct_options, index=0)
        amt2 = st.number_input("Amount 2", key="amt2")

    acct3 = st.selectbox("Account 3", options=acct_options, index=0)
    amt3 = st.number_input("Amount 3", key="amt3")

    if st.button("Submit Salary"):
        entries = []
        for acct, amt in [(acct1, amt1), (acct2, amt2), (acct3, amt3)]:
            if acct and amt > 0:
                uid = generate_unique_id(acct, amt)
                row = [
                    uid, salary_date, "Income", "Income", "", False,
                    "Salary / Income", acct, amt, amt, "USD", "Paycheck", None
                ]
                entries.append(row)
        if entries:
            add_transaction_to_excel(entries)
            st.success("Salary transactions added!")
